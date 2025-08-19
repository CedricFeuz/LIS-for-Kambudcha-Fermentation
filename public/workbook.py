# filepath: c:\Users\cedri\OneDrive - ZHAW\Studium\Semester 6\Laboratory Informatics\project\digital_lab_notebook\public\workbook.py
from openpyxl import load_workbook, Workbook
import os

# File path
file_path = 'users.xlsx'

# Check if the file exists
if not os.path.exists(file_path):
    # Create a new workbook if the file doesn't exist
    workbook = Workbook()
    workbook.save(file_path)
    print(f"File '{file_path}' created.")

# Load the existing workbook
workbook = load_workbook(file_path)

# Select the active worksheet
sheet = workbook.active

# Add headers for the columns
headers = ["name", "username", "password"]
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

# Add dummy data
dummy_data = [
    ["Alice Smith", "asmith", "password123"],
    ["Bob Johnson", "bjohnson", "securepass456"],
    ["Charlie Brown", "cbrown", "mypassword789"]
]

for row_num, row_data in enumerate(dummy_data, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook
workbook.save(file_path)
print("Dummy data added successfully.")
